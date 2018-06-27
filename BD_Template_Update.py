############################## BD Template update to multiple tabs of Automation sheet  ############################
#
#Description :  This script copies reference BD template to all the valid test case tabs of Automation Multiiteration Sheet 
#
# Revision History : 
# Release		Release_Date 		Author		            Description/Enchancements
#	V1.0		 26/06/2018	   	Jitendra Tomar	   	Designed and Developed by : jtomar@qti.qualcomm.com
#
####################################################################################################################

import sys
import openpyxl as xl

#Reference BD Template Path
BD_Template = sys.argv[1]
print ("Reference BD Template : ", BD_Template)

#Automation combined sheet with all Test Cases BD
Target_Sheet = sys.argv[2]
print ("Reference BD Template : ", Target_Sheet)

#Output Sheet
Out_Sheet = Target_Sheet
print("output sheet : ", Out_Sheet)

#Load Reference Sheet in openpyxl
wb_ref = xl.load_workbook(BD_Template)
print(wb_ref.worksheets[0])
ref_sheet = wb_ref.worksheets[0]

#Load Target Sheet in openpyxl
wb_target = xl.load_workbook(Target_Sheet)
print(wb_target.worksheets[0])

#tar_sheet = wb_target.worksheets[0]

#Function to update values in Target sheet from Reference sheet
def update(testcase_sheet):
    print("\nInside update function", testcase_sheet)
    for row in ref_sheet:
        for cell in row:
            testcase_sheet[cell.coordinate].value = cell.value

#Logic to read valid test case workbook sheet names
sheets_list = wb_target.sheetnames
sheet_no = 0

for TC_sheet in sheets_list:
    temp = TC_sheet.split()[-1]
    print(temp)
    print("sheet no : ", sheet_no)
    if temp == "(1)":
        print("Updating", TC_sheet, "values")
        update(wb_target.worksheets[sheet_no])
    elif temp == "(2)":
        print("Updating", TC_sheet, "values")
        update(wb_target.worksheets[sheet_no])
    elif temp == "(3)":
        print("Updating", TC_sheet, "values")
        update(wb_target.worksheets[sheet_no])
    else:
        print("\nNot valid sheet to update values", TC_sheet)
    sheet_no += 1

#update(tar_sheet)
wb_target.save(Out_Sheet)

print("\n======= Done Enjoy ==========")